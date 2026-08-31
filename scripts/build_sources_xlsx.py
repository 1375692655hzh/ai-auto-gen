# -*- coding: utf-8 -*-
"""从注册表生成源清单 Excel(global-news-sources/docs/源清单.xlsx)。数据=REGISTRY 实时导出+人工补充的作息/备注映射。
用法: py -3.11 scripts/build_sources_xlsx.py"""
import sys, io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "global-news-sources"))

XLSX_SKILL_DIR = r"C:\Users\hzh\.zcode\cli\plugins\cache\zcode-plugins-official\document-skills\0.1.4\skills\xlsx"
sys.path.insert(0, XLSX_SKILL_DIR)  # base.py 内部用 `from templates.palettes import` 需要 skill 根
sys.path.insert(0, os.path.join(XLSX_SKILL_DIR, "templates"))
import base  # noqa: E402
base.use_palette_explicit("bloomberg")  # 金融数据终端风格
from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Alignment  # noqa: E402

import sources.builtin  # noqa: E402,F401
from sources.base import REGISTRY  # noqa: E402

KIND_CN = {"flash": "快讯流", "announcement": "公告/披露", "market": "数据/宏观",
           "peer_article": "同行文章", "calendar": "日历", "peer_group": "聚合组",
           "extras_group": "素材组"}

# id -> (市场, 更新作息, 状态, 备注)
M = {
    # A股/中国
    "sina_7x24": ("A股/中国", "7×24 滚动", "启用", ""),
    "eastmoney_fast": ("A股/中国", "7×24 秒级", "启用", "焦点栏目"),
    "cls_telegraph": ("A股/中国", "7×24 实时", "启用", "本地签名零key"),
    "ths_flash": ("A股/中国", "7×24 滚动", "启用", "带官方重要性标记"),
    "caixin_flash": ("A股/中国", "每日滚动", "启用", "高质量中文财经, 带栏目标签"),
    "cninfo_latest": ("A股/中国", "按日期归集", "启用", "A股官方信披"),
    "cls_announcements": ("A股/中国", "交易时段", "素材·手动", "版式素材条目"),
    "nbs_pmi": ("A股/中国", "月度", "启用", "官方; 首页滚出后翻页兜底"),
    "pboc_social_financing": ("A股/中国", "月度", "启用", "官方, 总量+结构"),
    "cctv_xwlb": ("A股/中国", "每日19:00", "启用", "政策信号, 不入聚合, 供synth背景"),
    "cn_index_snapshot": ("A股/中国", "盘中实时", "启用", "market-review 复盘工作流用"),
    "em_sector_board": ("A股/中国", "交易日", "启用", "东财主+新浪备双风控面"),
    "em_research": ("A股/中国", "每日", "聚合", "机构观点索引"),
    "sina_vip": ("A股/中国", "日内", "聚合", "首席经济学家/大V观点"),
    "cls_morning": ("A股/中国", "每日07:00", "聚合", "有声早报"),
    "eastmoney_zaozhidao": ("A股/中国", "每早", "启用", "《早知道》系列"),
    "gangtise": ("A股/中国", "每日", "聚合", "搜狗微信链, 高风险"),
    "yuanbao_gangtise": ("A股/中国", "每日", "手动", "需 Playwright 登录态"),
    # 全球/多市场
    "jin10_flash": ("全球/多市场", "7×24 秒级", "启用", "全球宏观"),
    "eastmoney_global": ("全球/多市场", "日内滚动", "启用", "美/港/宏观, 与电报不同风控面"),
    "wscn_live": ("全球/多市场", "7×24", "启用", "含非财经条目, 下游粗筛"),
    "investinglive_flash": ("全球/多市场", "交易日滚动", "启用", "英文, FXStreet 同赛道顶替"),
    "fxstreet_flash": ("全球/多市场", "交易日滚动", "占位·被拦待恢复", "出口 IP 被 Cloudflare 整站 403"),
    "jin10_calendar": ("全球/多市场", "日更", "占位·被拦待恢复", "官方 CDN 域名 NXDOMAIN"),
    "frankfurter_fx": ("全球/多市场", "交易日一更", "启用", "ECB 参考汇率, 不含 TWD"),
    "goldprice_metals": ("全球/多市场", "小时级", "启用", "仅黄金"),
    "cftc_cot": ("全球/多市场", "周度", "启用", "官方 Socrata, 金银铜油/外汇/股指净多"),
    "smm_metals": ("全球/多市场", "每日", "聚合", "大宗商品日报"),
    "miningcom": ("全球/多市场", "日更", "启用", "大宗矿业新闻汇总"),
    "wscn_breakfast": ("全球/多市场", "每早", "启用", "华尔街见闻早餐"),
    "calendar": ("全球/多市场", "日更", "组合·手动", "旧版式素材"),
    "global_markets": ("全球/多市场", "盘中", "组合·手动", "旧版式素材"),
    "extras": ("全球/多市场", "—", "组合·手动", "版式素材聚合入口"),
    "peer_mornings": ("全球/多市场", "每早", "组合·手动", "17 源同行早报聚合入口"),
    # 美股
    "yahoo_headlines": ("美股", "交易日滚动", "启用", "大盘头条英文, 可换 symbol"),
    "marketwatch_rt": ("美股", "分钟级", "启用", "道琼斯 CDN; 旧 realtimeheadlines 死 feed 已弃"),
    "foxbusiness_rss": ("美股", "分钟级", "启用", ""),
    "benzinga_rss": ("美股", "分钟级", "启用", "夹加密/预测软文, 下游粗筛"),
    "finviz_news": ("美股", "秒~分钟", "启用", "聚合 Bloomberg/Reuters/WSJ/CNBC 等"),
    "seekingalpha_rt": ("美股", "分钟级", "启用", "带 ticker; feed 自述限非商用, 注意 ToS"),
    "sec_edgar": ("美股", "交易时段滚动", "启用", "官方 8-K 重大公告"),
    "sec_insider": ("美股", "交易时段滚动", "启用", "官方 Form4 内部人交易"),
    "sec_press": ("美股", "发布即更", "启用", "官方执法/规则制定"),
    "prnewswire": ("美股", "分钟级", "启用", "公司新闻稿原稿第一落点"),
    "globenewswire": ("美股", "分钟级", "占位·被拦待恢复", "本机出口 IP ReadTimeout, 疑 IDC 段限制"),
    "fed_press": ("美股", "事件驱动", "启用", "FOMC/监管执法, 最高权威源"),
    "treasury_press": ("美股", "发布即更", "启用", "仅 /rss.xml 有效路径"),
    "bea_rss": ("美股", "发布日 08:30 ET", "启用", "GDP/PCE; 须用 apps.bea.gov"),
    "finra_press": ("美股", "不定期", "启用", "执法/纪律处分; 必须走 http"),
    "nyfed_rates": ("美股", "交易日日更", "启用", "SOFR/EFFR/OBFR/TGCR 官方"),
    "bls_macro": ("美股", "月度发布日", "启用", "CPI/非农/失业率; v1 零 key 日限 25 次"),
    "fiscal_debt": ("美股", "工作日 T+1", "启用", "国债总额"),
    "gdpnow_rss": ("美股", "发布即推", "启用", "亚特兰大联储 GDPNow, 周 1-2 次"),
    "treasury_yield_curve": ("美股", "日更", "启用", "3M/2Y/10Y/30Y + 10Y-2Y 利差"),
    "nasdaq_earnings": ("美股", "交易日", "启用", "财报日历; 周末/假期空属正常"),
    "stocktwits_stream": ("美股", "交易时段高频", "占位·被拦待恢复", "Cloudflare 挑战拦截"),
    "reddit_hot": ("美股", "全天滚动", "占位·待配key", "免费 OAuth, reddit.com/prefs/apps 即申即得"),
    "polymarket_sentiment": ("美股", "实时", "启用", "预测市场财经概率+24h量"),
    "fred_macro": ("美股", "日更", "占位·待配key", "免费 key 即申即得"),
    "marketaux_news": ("美股", "滚动", "占位·待配key", "免费 ~100 次/天"),
    "finnhub_news": ("美股", "滚动", "占位·待配key", "免费 60 次/分"),
    "alphavantage_news": ("美股", "滚动", "占位·待配key", "免费仅 25 次/天, 应急"),
    "cnbc_morning": ("美股", "工作日晨", "聚合", "CNBC Daily Open"),
    "newsquawk_open": ("美股", "交易日午后/傍晚", "聚合", "欧美开盘综述"),
    "liberty_street": ("美股", "周 1-2 篇", "启用", "联储经济学家分析博客"),
    "eia_energy": ("美股", "工作日每日", "启用", "能源大宗基本面日报"),
    # 港股
    "hkexnews": ("港股", "实时随发随更", "启用", "披露易法定披露一手源, 代码+PDF 直链"),
    "mingpao_rss": ("港股", "盘中高频", "启用", "IPO 聆讯/中报/本地宏观"),
    "yahoo_hk_rss": ("港股", "分钟级", "启用", "AASTOCKS AAFN 内容, 繁体带代码"),
    "scmp_biz_rss": ("港股", "日内滚动", "启用", "英文; 标题摘要免费, 正文软付费墙"),
    "eastmoney_hkus": ("港股", "7×24 秒级", "启用", "fastColumn=104 港美股频道"),
    "rthk_finance": ("港股", "近 7×24", "启用", "周末有 ADR/金油汇"),
    "hkex_press": ("港股", "不定期", "启用", "港交所自身新闻稿(规则/产品)"),
    "sfc_press": ("港股", "工作日不定期", "启用", "证监会执法/互联互通"),
    "longbridge_topics": ("港股", "日内", "启用", "港美券商话题, TanStack 脱水 JSON 解析"),
    "gelonghui": ("港股", "日更", "启用", "港/美/A 评论要闻"),
    "futu_morning": ("港股", "工作日晨", "聚合", "富途《港美早报》"),
    "etnet_open": ("港股", "工作日 08:30", "聚合", "etnet 開市 Go"),
    "hkma_press": ("港股", "发布即更", "占位·被拦待恢复", "本机出口 TLS 被重置, 待部署环境复测"),
    "hkgov_finance": ("港股", "不定期", "占位·被拦待恢复", "本机出口 TLS 被重置, 待部署环境复测"),
    # 台湾
    "cna_flash": ("台湾", "日内高频", "启用", "官方通讯社 JSON API, 单页 100 条"),
    "twse_news": ("台湾", "交易日", "启用", "证交所新闻+法说会日历, 民国年已转公元"),
    "twse_mops": ("台湾", "日更出表", "启用", "上市公司每日重大讯息"),
    "tpex_mops": ("台湾", "日更出表", "启用", "上柜公司每日重大讯息"),
    "yahoo_tw_rss": ("台湾", "分钟级 ttl=5", "启用", "可按个股 ?s=2330.TW 订阅"),
    "udn_rss": ("台湾", "日内高频", "启用", "经济日报; 含非财经条目, 下游粗筛"),
    "ltn_rss": ("台湾", "日内滚动", "启用", "自由时报财经"),
    "technews_rss": ("台湾", "日内滚动", "启用", "半导体/供应链, 台积电链素材"),
    "moneydj_flash": ("台湾", "日内滚动", "启用", "盘口快讯; RSS 已退化空壳走 HTML 解析"),
    "cnyes_tw": ("台湾", "日内(周末照常)", "聚合", "鉅亨网台股精选"),
    "threads_kol_digest": ("台湾", "每日(跟随监控项目)", "启用", "Threads KOL 情报日报; 本地文件源(threads-tw-monitor 产物), 可配 digests_dir"),
    "tw_cbc_stats": ("台湾", "按月", "占位·被拦待恢复", "央行统计 API; 本机出口 TLS 被重置"),
    "fsc_press": ("台湾", "工作日", "占位·被拦待恢复", "金管会新闻稿; 本机出口 TLS 被重置"),
    "finmind_news": ("台湾", "滚动", "占位·待配key", "免费 token 即申即得"),
    # 土耳其
    "kap_disclosures": ("土耳其", "近实时(周末亦有)", "启用", "法定公告一手源; 必须 POST JSON"),
    "tcmb_fx": ("土耳其", "交易日 15:30 TRT", "启用", "央行官方汇率牌价 XML"),
    "tcmb_press": ("土耳其", "事件驱动", "启用", "利率决议/流动性操作, Atom 格式"),
    "dailysabah_rss": ("土耳其", "日内高频", "启用", "英文商业新闻"),
    "hurriyet_rss": ("土耳其", "日内多次", "启用", "英文; business 频道已死, 接 /rss/news"),
    "dunya_rss": ("土耳其", "5 分钟刷新", "启用", "土语主流财经日报"),
    "cnbce_rss": ("土耳其", "分钟级", "启用", "土语, 单次约 250 条"),
    "foreks_rss": ("土耳其", "小时级", "启用", "土语, 本土财经数据商"),
    "sabah_rss": ("土耳其", "日更约 10 条", "启用", "宏观日程/政策, 早报语境"),
    "aa_morning": ("土耳其", "每日晨", "聚合", "AA 通讯社英文晨报"),
    "turkey_morning": ("土耳其", "交易日晨", "聚合", "BloombergHT"),
    "tcmb_evds": ("土耳其", "按指标发布", "占位·待配key", "央行 EVDS2 宏观序列, 免费 key 即申即得"),
    # 日韩
    "japan_morning": ("日本", "日内 RSS 过滤当日", "聚合", "共同社"),
    "korea_morning": ("韩国", "日内 RSS 过滤当日", "聚合", "韩联社"),
}

rows = []
missing = [k for k in REGISTRY if k not in M]
assert not missing, f"映射缺: {missing}"
extra = [k for k in M if k not in REGISTRY]
assert not extra, f"映射多出: {extra}"
for sid, v in REGISTRY.items():
    meta = v["meta"]
    market, schedule, status, note = M[sid]
    on = meta["default_enabled"]
    if on:
        assert status == "启用", f"{sid} 注册为启用但映射状态={status}"
    else:
        assert status != "启用", f"{sid} 注册为停用但映射状态=启用"
    rows.append([len(rows) + 1, sid, market, KIND_CN[meta["kind"]],
                 meta["title"], schedule, status, note])
assert len(rows) == 108, len(rows)

wb = Workbook()

# ============ Sheet 1: 源清单 ============
ws = wb.active
ws.title = "源清单"
headers = ["序号", "源 id", "市场", "类型", "给什么内容", "更新作息", "状态", "备注/坑"]
last_col = len(headers) + 1  # B..I
base.setup_sheet(ws, title="ai-auto-gen 信息源清单（注册表全量 108 源，2026-08-31 导出）",
                 last_col=last_col)
for c, h in enumerate(headers, start=2):
    ws.cell(row=4, column=c, value=h)
base.style_header_row(ws, row_num=4, col_start=2, col_end=last_col)
for i, r in enumerate(rows):
    for c, v in enumerate(r, start=2):
        ws.cell(row=5 + i, column=c, value=v)
    base.style_data_row(ws, row_num=5 + i, col_start=2, col_end=last_col, row_index=i)
    ws.cell(row=5 + i, column=2).alignment = Alignment(horizontal="center", vertical="center")
ws.freeze_panes = "C5"
base.auto_fit_columns(ws, min_width=8, max_width=40, header_row=4, data_start_row=5)
base.auto_fit_row_heights(ws, header_row=4, data_start_row=5)
ws.page_setup.orientation = "landscape"
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.print_title_rows = "4:4"

wb.save(r"E:\ai-gen-article-publish\global-news-sources\docs\源清单.xlsx")
print("sheet1 done,", len(rows), "rows")

# ============ Sheet 2: 总览 ============
ws2 = wb.create_sheet("总览")
base.setup_sheet(ws2, title="总览：按市场与状态统计（活公式，随源清单联动）", last_col=7)

# 表1: 市场 × 状态
h2 = ["市场", "总数", "启用", "占位·被拦待恢复", "占位·待配key", "聚合/手动"]
for c, h in enumerate(h2, start=2):
    ws2.cell(row=4, column=c, value=h)
base.style_header_row(ws2, row_num=4, col_start=2, col_end=7)
markets = ["A股/中国", "全球/多市场", "美股", "港股", "台湾", "土耳其", "日本", "韩国"]
list_range = "'源清单'!$D$5:$D$111"   # 市场列
stat_range = "'源清单'!$H$5:$H$111"   # 状态列
for i, mk in enumerate(markets):
    r = 5 + i
    ws2.cell(row=r, column=2, value=mk)
    ws2.cell(row=r, column=3, value=f"=COUNTIF({list_range},B{r})")
    ws2.cell(row=r, column=4, value=f'=COUNTIFS({list_range},B{r},{stat_range},"启用")')
    ws2.cell(row=r, column=5, value=f'=COUNTIFS({list_range},B{r},{stat_range},"占位·被拦待恢复")')
    ws2.cell(row=r, column=6, value=f'=COUNTIFS({list_range},B{r},{stat_range},"占位·待配key")')
    ws2.cell(row=r, column=7, value=f'=COUNTIFS({list_range},B{r},{stat_range},"聚合")'
                                    f'+COUNTIFS({list_range},B{r},{stat_range},"组合·手动")'
                                    f'+COUNTIFS({list_range},B{r},{stat_range},"素材·手动")'
                                    f'+COUNTIFS({list_range},B{r},{stat_range},"手动")')
    base.style_data_row(ws2, row_num=r, col_start=2, col_end=7, row_index=i)
tot_r = 5 + len(markets)
ws2.cell(row=tot_r, column=2, value="合计")
for c in range(3, 8):
    col = base.get_column_letter(c) if hasattr(base, "get_column_letter") else None
from openpyxl.utils import get_column_letter
for c in range(3, 8):
    L = get_column_letter(c)
    ws2.cell(row=tot_r, column=c, value=f"=SUM({L}5:{L}{tot_r - 1})")
base.style_total_row(ws2, row_num=tot_r, col_start=2, col_end=7)

# 表2: 类型分布
t2_start = tot_r + 3
ws2.cell(row=t2_start, column=2, value="按类型").font = base.font_subheader()
h3 = ["类型", "数量"]
for c, h in enumerate(h3, start=2):
    ws2.cell(row=t2_start + 1, column=c, value=h)
base.style_header_row(ws2, row_num=t2_start + 1, col_start=2, col_end=3)
kind_range = "'源清单'!$E$5:$E$111"
kinds = ["快讯流", "公告/披露", "数据/宏观", "同行文章", "日历", "聚合组", "素材组"]
for i, kd in enumerate(kinds):
    r = t2_start + 2 + i
    ws2.cell(row=r, column=2, value=kd)
    ws2.cell(row=r, column=3, value=f'=COUNTIF({kind_range},B{r})')
    base.style_data_row(ws2, row_num=r, col_start=2, col_end=3, row_index=i)

# 说明
note_r = t2_start + 2 + len(kinds) + 2
notes = [
    "口径：sources/base.py REGISTRY 实时导出（python cli.py sources list 同源），作息/备注为人工维护映射。",
    "状态释义：启用=默认进采集池；聚合=经 peer_mornings 同行早报聚合采集（单行默认关）；占位=代码在册默认关，被拦待恢复或待配免费 key。",
    "法定披露层五市场齐备：A股巨潮 / 港股披露易 / 美股 EDGAR / 台湾 TWSE+TPEx 重大讯息 / 土耳其 KAP。",
    "生成时选取：flows run --set flash_sources=... 或 config.yaml 的 sources.<id>.enabled。",
]
for i, t in enumerate(notes):
    cell = ws2.cell(row=note_r + i, column=2, value=t)
    cell.font = base.font_caption()
    cell.alignment = Alignment(horizontal="left", vertical="center")

base.auto_fit_columns(ws2, min_width=8, max_width=40, header_row=4, data_start_row=5)
base.auto_fit_row_heights(ws2, header_row=4, data_start_row=5)

wb.properties.creator = "Z.ai"
wb.save(r"E:\ai-gen-article-publish\global-news-sources\docs\源清单.xlsx")
print("saved global-news-sources/docs/源清单.xlsx")
